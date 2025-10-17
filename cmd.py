from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

def add_data_with_labels(ws, title, start_row, data, year):
    total = round(sum(v for _, v in data), 1)
    ws.cell(row=start_row, column=1, value=f"{year} - {title}")
    ws.cell(row=start_row+1, column=1, value="Ngành")
    ws.cell(row=start_row+1, column=2, value="Giá trị")
    ws.cell(row=start_row+1, column=3, value="Tỉ lệ (%)")
    ws.cell(row=start_row+1, column=4, value="Nhãn")
    r = start_row + 2
    for sector, value in data:
        percent = value / total * 100
        label = f"{percent:.1f}% ({value})"
        ws.cell(row=r, column=1, value=sector)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=round(percent, 1))
        ws.cell(row=r, column=4, value=label)
        r += 1
    return total, start_row+2, r-1

wb = Workbook()
ws = wb.active
ws.title = "2010-2021"
data_2010 = [("Nông nghiệp", 675.4), ("Lâm nghiệp", 22.8), ("Thủy sản", 177.7)]
data_2021 = [("Nông nghiệp", 1502.2), ("Lâm nghiệp", 63.3), ("Thủy sản", 559.7)]
total_2010, r1_start, r1_end = add_data_with_labels(ws, "Cơ cấu ngành", 1, data_2010, 2010)
total_2021, r2_start, r2_end = add_data_with_labels(ws, "Cơ cấu ngành", 10, data_2021, 2021)
pie1 = PieChart()
data = Reference(ws, min_col=2, min_row=r1_start, max_row=r1_end)
labels = Reference(ws, min_col=4, min_row=r1_start, max_row=r1_end)
pie1.add_data(data, titles_from_data=False)
pie1.set_categories(labels)
pie1.title = f"2010 (Tổng: {round(total_2010)})"
ws.add_chart(pie1, "F2")
pie2 = PieChart()
data2 = Reference(ws, min_col=2, min_row=r2_start, max_row=r2_end)
labels2 = Reference(ws, min_col=4, min_row=r2_start, max_row=r2_end)
pie2.add_data(data2, titles_from_data=False)
pie2.set_categories(labels2)
pie2.title = f"2021 (Tổng: {round(total_2021)})"
ws.add_chart(pie2, "F20")
file_path_combined = "./nd2.xlsx"
wb.save(file_path_combined)

file_path_combined
